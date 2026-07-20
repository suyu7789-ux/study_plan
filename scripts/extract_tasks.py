import json
import argparse
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def date_text(value):
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value or "")[:10]


parser = argparse.ArgumentParser(description="Extract categorized tasks from the study plan workbook")
parser.add_argument("source", type=Path, help="Path to the source .xlsx workbook")
parser.add_argument(
    "--output",
    type=Path,
    default=ROOT / "app" / "tasks.json",
    help="Destination JSON path",
)
args = parser.parse_args()

workbook = load_workbook(args.source, read_only=True, data_only=True)
sheet = workbook["分类任务"]
tasks = []

for row in sheet.iter_rows(min_row=5, max_row=844, values_only=True):
    if not row[0]:
        continue
    tasks.append(
        {
            "id": row[0],
            "date": date_text(row[1]),
            "day": int(str(row[2]).replace("Day ", "")),
            "week": int(str(row[3]).replace("第", "").replace("周", "")),
            "subject": row[4],
            "category": row[5],
            "source": row[6],
            "section": row[7],
            "detail": row[8],
            "output_standard": row[9],
            "planned_minutes": int(row[10] or 0),
            "min_images": int(row[12] or 0),
        }
    )

args.output.parent.mkdir(parents=True, exist_ok=True)
args.output.write_text(json.dumps(tasks, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {len(tasks)} tasks to {args.output}")
